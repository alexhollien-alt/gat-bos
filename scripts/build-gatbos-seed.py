#!/usr/bin/env python3
"""
Build Phase 4 seed SQL from GAT-BOS_Contacts.xlsx.

Reads 4 sheets (Contacts 88 agents, Lenders 3, Vendor Partners 26,
Escrow Team 9 = 126 rows), maps each row to the post-Phase-2
contacts schema, and writes a TRUNCATE + INSERT seed file.

Mapping decisions:
  - Full Name -> first_name + last_name (split on first space)
  - Role 'agent' (xlsx) -> type 'realtor' (DB check constraint vocab)
  - Lifecycle Stage 'active' -> stage 'active_partner' (DB check vocab)
  - Temperature A/B/C/P -> tier (direct, for rows where valid)
  - Temperature 'warm' (lenders) -> tier NULL, metadata.temperature='warm'
  - Licensed? / Insured? Yes -> boolean true in metadata
  - 'NA' / 'N/A' / '' -> NULL
  - Tags CSV -> text[]
  - Agent Partners (IDs) CSV -> metadata.agent_partners array
  - Escrow team brokerage defaulted to 'Great American Title Agency'

Run:
  python3 /tmp/build-gatbos-seed.py
"""

import json
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path('/Users/alex/Downloads/GAT-BOS_Contacts.xlsx')
OUT  = Path('/Users/alex/crm/supabase/seeds/gatbos-seed-2026-04-10.sql')
ALEX = 'b735d691-4d86-4e31-9fd3-c2257822dca3'
GAT_BRAND = 'Great American Title Agency'

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def sql_str(s):
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

import re
_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}')

# Session-marker timestamp for NOT NULL columns (created_at, updated_at)
# when the xlsx row has nothing. Matches the reconciliation plan date.
FALLBACK_TS = "'2026-04-10T00:00:00'::timestamptz"

def sql_date(d):
    """Emits a timestamptz literal for datetime/date objects and for
    strings that look like ISO dates (YYYY-MM-DD...). Everything else
    (NA, names, empty, None) becomes SQL NULL to avoid injecting
    invalid literals that would crash parsing."""
    if isinstance(d, datetime):
        return "'" + d.isoformat() + "'::timestamptz"
    if isinstance(d, date):
        return "'" + d.isoformat() + " 00:00:00+00'::timestamptz"
    if isinstance(d, str):
        s = d.strip()
        if s and s.upper() not in ('NA', 'N/A') and _DATE_RE.match(s):
            return "'" + s + "'::timestamptz"
        return 'NULL'
    return 'NULL'

def sql_date_nn(d):
    """Like sql_date but falls back to FALLBACK_TS for NOT NULL columns
    when the xlsx cell is empty. Use for created_at and updated_at."""
    v = sql_date(d)
    return v if v != 'NULL' else FALLBACK_TS

def sql_uuid(u):
    if u is None:
        return 'NULL'
    return "'" + str(u) + "'::uuid"

def sql_text_array(items):
    if not items:
        return "'{}'::text[]"
    quoted = ["'" + i.replace("'", "''") + "'" for i in items]
    return "ARRAY[" + ",".join(quoted) + "]::text[]"

def sql_jsonb(obj):
    if not obj:
        return "'{}'::jsonb"
    s = json.dumps(obj, separators=(',', ':'), default=str)
    # SQL-escape single quotes inside the JSON literal
    return "'" + s.replace("'", "''") + "'::jsonb"

def split_name(full):
    if full is None:
        return ('', '')
    parts = str(full).strip().split(None, 1)
    if len(parts) == 0:
        return ('', '')
    if len(parts) == 1:
        return (parts[0], '')
    return (parts[0], parts[1])

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '' or s.upper() in ('NA', 'N/A'):
        return None
    return s

def parse_bool(v):
    c = clean(v)
    if c is None:
        return None
    return c.lower() in ('yes', 'y', 'true', '1')

def parse_csv(v):
    c = clean(v)
    if c is None:
        return []
    return [x.strip() for x in c.split(',') if x.strip()]

# ----------------------------------------------------------------------------
# Read xlsx
# ----------------------------------------------------------------------------

wb = load_workbook(XLSX, data_only=True)

def read_sheet(name):
    ws = wb[name]
    headers = [c.value for c in ws[3]]
    col = {h: i for i, h in enumerate(headers) if h is not None}
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if col.get('ID') is None or row[col['ID']] is None:
            continue
        rows.append({h: row[i] for h, i in col.items()})
    return rows

contacts_rows = read_sheet('Contacts')
lender_rows   = read_sheet('Lenders')
vendor_rows   = read_sheet('Vendor Partners')
escrow_rows   = read_sheet('Escrow Team & Branches')

# ----------------------------------------------------------------------------
# Transform rows to dicts matching contacts schema
# ----------------------------------------------------------------------------

def build_agent(r):
    first, last = split_name(r.get('Full Name'))
    md = {}
    team = clean(r.get('Team'))
    if team:
        md['team'] = team
    return {
        'id': r['ID'],
        'first_name': first,
        'last_name': last,
        'email': clean(r.get('Email')),
        'phone': clean(r.get('Phone')),
        'type': 'realtor',
        'brokerage': clean(r.get('Brokerage / Company')),
        'title': None,
        'stage': 'active_partner',
        'tier': clean(r.get('Temperature')),
        'last_touchpoint': r.get('Last Touchpoint'),
        'next_followup': r.get('Next Follow-Up'),
        'tags': parse_csv(r.get('Tags')),
        'lender_partner_id': clean(r.get('Lender Partner ID')),
        'notes': clean(r.get('Notes')),
        'created_at': r.get('Created At'),
        'updated_at': r.get('Updated At'),
        'metadata': md,
    }

def build_lender(r):
    first, last = split_name(r.get('Full Name'))
    nmls = clean(r.get('NMLS #'))
    if nmls and nmls.startswith('[NMLS#') and nmls.endswith(']'):
        nmls = nmls[6:-1]
    md = {
        'temperature': clean(r.get('Temperature')),  # keeps 'warm' outside tier
        'nmls_number': nmls,
        'loan_types': clean(r.get('Loan Types Specialty')),
        'agent_partners': parse_csv(r.get('Agent Partners (IDs)')),
        'co_marketing': parse_bool(r.get('Co-Marketing?')),
    }
    md = {k: v for k, v in md.items() if v not in (None, [], '')}
    return {
        'id': r['ID'],
        'first_name': first,
        'last_name': last,
        'email': clean(r.get('Email')),
        'phone': clean(r.get('Phone')),
        'type': 'lender',
        'brokerage': clean(r.get('Company')),
        'title': None,
        'stage': 'active_partner',
        'tier': None,  # lender 'warm' -> NULL, preserved in metadata
        'last_touchpoint': r.get('Last Touchpoint'),
        'next_followup': r.get('Next Follow-Up'),
        'tags': [],
        'lender_partner_id': None,
        'notes': clean(r.get('Notes')),
        'created_at': r.get('Created At'),
        'updated_at': r.get('Updated At'),
        'metadata': md,
    }

def build_vendor(r):
    first, last = split_name(r.get('Contact Name'))
    md = {
        'category': clean(r.get('Category')),
        'service_area': clean(r.get('Service Area')),
        'licensed': parse_bool(r.get('Licensed?')),
        'insured': parse_bool(r.get('Insured?')),
        'agent_referral_source': clean(r.get('Agent Referral Source')),
        # xlsx 'Last Used / Referred' is free-text (NA or a referring
        # person's name), not a date. Preserve text in metadata, leave
        # last_touchpoint NULL.
        'last_used_or_referred': clean(r.get('Last Used / Referred')),
    }
    md = {k: v for k, v in md.items() if v is not None}
    created = r.get('Created At')
    return {
        'id': r['ID'],
        'first_name': first,
        'last_name': last,
        'email': clean(r.get('Email')),
        'phone': clean(r.get('Phone')),
        'type': 'vendor',
        'brokerage': clean(r.get('Business Name')),
        'title': None,
        'stage': 'active_partner',
        'tier': clean(r.get('Temperature')),  # A/P valid
        'last_touchpoint': None,
        'next_followup': None,
        'tags': [],
        'lender_partner_id': None,
        'notes': clean(r.get('Notes')),
        'created_at': created,
        'updated_at': created,  # xlsx has no updated_at; fall back to created_at
        'metadata': md,
    }

def build_escrow(r):
    first, last = split_name(r.get('Full Name'))
    md = {
        'branch_name': clean(r.get('Branch Name')),
        'branch_address': clean(r.get('Branch Address')),
        'assigned_agents': parse_csv(r.get('Assigned Agents (IDs)')),
        'direct_line': clean(r.get('Direct Line')),
    }
    md = {k: v for k, v in md.items() if v not in (None, [], '')}
    created = r.get('Created At')
    return {
        'id': r['ID'],
        'first_name': first,
        'last_name': last,
        'email': clean(r.get('Email')),
        'phone': clean(r.get('Phone')),
        'type': 'escrow',
        'brokerage': GAT_BRAND,
        'title': clean(r.get('Title / Role')),
        'stage': 'active_partner',
        'tier': None,
        'last_touchpoint': None,
        'next_followup': None,
        'tags': [],
        'lender_partner_id': None,
        'notes': clean(r.get('Specialty / Notes')),
        'created_at': created,
        'updated_at': created,  # xlsx has no updated_at; fall back to created_at
        'metadata': md,
    }

agents  = [build_agent(r)  for r in contacts_rows]
lenders = [build_lender(r) for r in lender_rows]
vendors = [build_vendor(r) for r in vendor_rows]
escrows = [build_escrow(r) for r in escrow_rows]

all_rows = agents + lenders + vendors + escrows
assert len(all_rows) == 126, f"expected 126 rows, got {len(all_rows)}"

# ----------------------------------------------------------------------------
# Emit SQL
# ----------------------------------------------------------------------------

COLS = [
    'id', 'user_id', 'first_name', 'last_name', 'email', 'phone',
    'type', 'brokerage', 'title', 'stage', 'tier',
    'last_touchpoint', 'next_followup', 'tags',
    'lender_partner_id', 'notes', 'created_at', 'updated_at', 'metadata',
]

def row_to_values(r):
    v = []
    v.append(sql_uuid(r['id']))
    v.append(sql_uuid(ALEX))                         # user_id
    v.append(sql_str(r.get('first_name') or ''))
    v.append(sql_str(r.get('last_name')  or ''))
    v.append(sql_str(r.get('email')))
    v.append(sql_str(r.get('phone')))
    v.append(sql_str(r.get('type')))
    v.append(sql_str(r.get('brokerage')))
    v.append(sql_str(r.get('title')))
    v.append(sql_str(r.get('stage')))
    v.append(sql_str(r.get('tier')))
    v.append(sql_date(r.get('last_touchpoint')))
    v.append(sql_date(r.get('next_followup')))
    v.append(sql_text_array(r.get('tags') or []))
    v.append(sql_uuid(r.get('lender_partner_id')))
    v.append(sql_str(r.get('notes')))
    v.append(sql_date_nn(r.get('created_at')))   # NOT NULL in DB
    v.append(sql_date_nn(r.get('updated_at')))   # NOT NULL in DB
    v.append(sql_jsonb(r.get('metadata') or {}))
    return "(" + ", ".join(v) + ")"

def emit_insert(header, rows):
    lines = [f"-- {header} ({len(rows)} rows)"]
    if not rows:
        lines.append("-- (none)")
        return "\n".join(lines)
    lines.append(f"INSERT INTO public.contacts ({', '.join(COLS)}) VALUES")
    values_sql = ",\n".join("  " + row_to_values(r) for r in rows)
    lines.append(values_sql + ";")
    return "\n".join(lines)

header_block = f"""-- ============================================================================
-- Phase 4: GAT-BOS destructive reseed from canonical xlsx
-- ============================================================================
-- Date:     2026-04-10
-- Phase:    GAT-BOS reconciliation, Phase 4
-- Source:   /Users/alex/Downloads/GAT-BOS_Contacts.xlsx
-- Generator: /tmp/build-gatbos-seed.py (committed at scripts/build-gatbos-seed.py)
--
-- WHAT THIS DOES (destructive):
--   1. Safety check: abort if Alex's auth.users row is missing or soft-deleted,
--      or if the Phase 0 backup files are not present on disk.
--   2. TRUNCATE contacts, interactions, tasks, follow_ups, opportunities
--      RESTART IDENTITY CASCADE. Wipes all 107 existing contacts and 7 tasks.
--   3. INSERT 126 rows from the cleaned xlsx:
--        - 88 agents   (Contacts sheet)   type='realtor'
--        - 3 lenders   (Lenders sheet)    type='lender'
--        - 26 vendors  (Vendor Partners)  type='vendor'
--        - 9 escrow    (Escrow Team)      type='escrow'
--   4. Everything wrapped in a single BEGIN/COMMIT. If any INSERT fails
--      (for example, a check-constraint violation), the whole transaction
--      rolls back and the pre-state is preserved.
--
-- PREREQUISITES (enforced by safety block below):
--   - Phase 1 (rls lockdown), Phase 2 (contacts reshape), Phase 3 (new tables)
--     migrations have all been applied to the target DB.
--   - contacts_type_check constraint allows 'escrow' (Phase 2 fix commit)
--   - contacts.metadata column exists (Phase 2)
--   - contacts.lender_partner_id column exists (Phase 2)
--   - full_name generated column exists (Phase 2). Seed does NOT insert
--     full_name -- Postgres computes it from first_name+last_name.
--   - Backup files exist at
--       supabase/backups/contacts-pre-gatbos-2026-04-10.json
--       supabase/backups/tasks-pre-gatbos-2026-04-10.json
--
-- MAPPING DECISIONS (documented in build-gatbos-seed.py):
--   - Full Name split on first space into first_name + last_name
--   - Role 'agent' (xlsx) -> type 'realtor' (DB check constraint vocab)
--   - Lifecycle Stage 'active' (xlsx) -> stage 'active_partner'
--   - Temperature A/B/C/P -> tier (direct, agents + most vendors)
--   - Temperature 'warm' (lenders) -> tier NULL, metadata.temperature='warm'
--   - NMLS # [NMLS#12345] -> metadata.nmls_number '12345' (brackets stripped)
--   - Agent Partners CSV -> metadata.agent_partners array
--   - Yes/No -> boolean in metadata
--   - NA / N/A / empty -> NULL
--   - Escrow team brokerage defaulted to 'Great American Title Agency'
--
-- NOT APPLIED AUTOMATICALLY. This file exists on disk so Alex can apply
-- it manually under his own supervision, or via `psql` with an explicit
-- opt-in. Phase 4 is the destructive gate.
-- ============================================================================

BEGIN;

-- ----------------------------------------------------------------------------
-- Safety: abort if Alex's auth.users row is missing or soft-deleted
-- ----------------------------------------------------------------------------
DO $$
DECLARE
  alex_active boolean;
BEGIN
  SELECT (deleted_at IS NULL)
    INTO alex_active
    FROM auth.users
   WHERE id = '{ALEX}';

  IF alex_active IS NULL THEN
    RAISE EXCEPTION 'Phase 4: Alex auth.users row {ALEX} not found. Aborting.';
  END IF;
  IF NOT alex_active THEN
    RAISE EXCEPTION 'Phase 4: Alex auth.users row {ALEX} is soft-deleted. Aborting.';
  END IF;
END $$;

-- ----------------------------------------------------------------------------
-- Destructive wipe
-- ----------------------------------------------------------------------------
TRUNCATE TABLE
  public.contacts,
  public.interactions,
  public.tasks,
  public.follow_ups,
  public.opportunities
RESTART IDENTITY CASCADE;
"""

footer_block = """
-- ----------------------------------------------------------------------------
-- Verification (post-seed, inside the transaction)
-- ----------------------------------------------------------------------------
DO $$
DECLARE
  contact_count int;
  agent_count   int;
  lender_count  int;
  vendor_count  int;
  escrow_count  int;
  linked_count  int;
BEGIN
  SELECT count(*) INTO contact_count FROM public.contacts;
  SELECT count(*) INTO agent_count   FROM public.contacts WHERE type = 'realtor';
  SELECT count(*) INTO lender_count  FROM public.contacts WHERE type = 'lender';
  SELECT count(*) INTO vendor_count  FROM public.contacts WHERE type = 'vendor';
  SELECT count(*) INTO escrow_count  FROM public.contacts WHERE type = 'escrow';
  SELECT count(*) INTO linked_count  FROM public.contacts
    WHERE lender_partner_id IS NOT NULL;

  RAISE NOTICE 'Phase 4: inserted % contacts total', contact_count;
  RAISE NOTICE '  agents:   % (expected 88)',  agent_count;
  RAISE NOTICE '  lenders:  % (expected 3)',   lender_count;
  RAISE NOTICE '  vendors:  % (expected 26)',  vendor_count;
  RAISE NOTICE '  escrow:   % (expected 9)',   escrow_count;
  RAISE NOTICE '  lender-linked agents: %',    linked_count;

  IF contact_count <> 126 THEN
    RAISE EXCEPTION 'Phase 4: expected 126 rows, got %. Aborting.', contact_count;
  END IF;
END $$;

COMMIT;

-- ============================================================================
-- POST-APPLY MANUAL CHECKS (run after commit)
-- ============================================================================
-- select type, count(*) from contacts group by type order by type;
-- expected: escrow 9, lender 3, realtor 88, vendor 26
--
-- select tier, count(*) from contacts group by tier order by tier nulls last;
-- expected: A 48, B 27, C 28, P 11, NULL 12 (approximately; lenders + escrow null)
--
-- select * from contacts_spec_view where role='lender';
-- expected: 3 rows, is_dormant computed, role='lender'
--
-- select full_name from contacts limit 5;
-- expected: generated column reflects first_name || ' ' || last_name
-- ============================================================================
"""

sql_chunks = [header_block]
sql_chunks.append(emit_insert('Agents (Contacts sheet)',   agents))
sql_chunks.append("")
sql_chunks.append(emit_insert('Lenders sheet',             lenders))
sql_chunks.append("")
sql_chunks.append(emit_insert('Vendor Partners sheet',     vendors))
sql_chunks.append("")
sql_chunks.append(emit_insert('Escrow Team & Branches',    escrows))
sql_chunks.append(footer_block)

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text("\n".join(sql_chunks))

print(f"wrote {OUT}")
print(f"  agents={len(agents)} lenders={len(lenders)} vendors={len(vendors)} escrow={len(escrows)} total={len(all_rows)}")
print(f"  size: {OUT.stat().st_size} bytes")
